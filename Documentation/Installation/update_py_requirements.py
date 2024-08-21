# This Python file uses the following encoding: utf-8
"""
*****************************************************************************
 @file    update_py_requirements.py
 @brief   Update version in requirement files
*****************************************************************************
"""

# autopep8: off
import sys
import os
import logging
import re
from typing import NamedTuple, List
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import table
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border

sys.path.append(os.path.join(os.path.dirname(__file__), "../../"))
from Source.Util.colored_log import init_console_logging  # pylint: disable=wrong-import-position
# autopep8: on

log = logging.getLogger("UpdatePyRequirements")
init_console_logging(logging.INFO)

L_FILES = ["requirements.txt", "constraints.txt"]  # files to update
L_IGNORE_PACKAGES = []
MAX_COLUMN_WIDTH = 30
FONT_SIZE = 11
FONT_NAME = "Consolas"
I_TIMEOUT = 2  # timeout to get pip data

L_PACKAGE_INFO_TITLE = ["Package", "Version", "Author", "Author_Mail", "License", "Homepage", "Package URL", "Requires", "Requires Py"]


class PackageInfo(NamedTuple):
    """!
    @brief Package Information.
    """
    package: str
    version: int
    author: str
    author_mail: str
    license_info: str
    home_page: str
    package_url: str
    requires_dist: str
    requires_py: str


def set_column_autowidth(worksheet: Worksheet, b_limit: bool = True):
    """!
    @brief Set automatic column width of worksheet.
    @param worksheet : select worksheet
    @param b_limit : status if width has a max limit
    """
    for i, col_cells in enumerate(worksheet.columns, start=1):
        i_max_col_len = 0
        for j, cell in enumerate(col_cells):
            if b_limit:
                if j != 0:  # do use not use first line description
                    i_max_col_len = max(i_max_col_len, len(str(cell.value).split("n", maxsplit=1)[0]))
            else:
                for s_line in str(cell.value).split("/n"):
                    i_max_col_len = max(i_max_col_len, len(s_line))
        if b_limit:
            i_max_col_len = min(i_max_col_len, MAX_COLUMN_WIDTH)
        if i_max_col_len == 0:
            worksheet.column_dimensions[get_column_letter(i)].hidden = True  # hide empty lines
        else:
            worksheet.column_dimensions[get_column_letter(i)].width = (i_max_col_len + 1) * 0.10 * FONT_SIZE


def set_cell(ws: Worksheet, i_row: int, i_column: int, value: any = None, b_bold: bool = False, b_italic: bool = False, b_underline: bool = False,
             i_font_size: int = 12, s_font: str = FONT_NAME, fill_color: str = None, align: str = None, s_format: str = None, s_border: Border = None):
    """!
    @brief Set cell data
    @param ws : actual worksheet
    @param i_row : row position
    @param i_column : column position
    @param value : value to set in cell; None: set no data to cell
    @param b_bold : status if cell content should be bold
    @param b_italic : status if cell content should be italic
    @param b_underline : status if cell content should be underlined
    @param i_font_size : font size
    @param s_font : font art
    @param fill_color : background fill color of cell
    @param align : text align of cell
    @param s_format : format of cell
    @param s_border : boarder of cell
    """
    cell = ws.cell(row=i_row, column=i_column)
    if value is not None:
        cell.value = value
    if b_underline:
        s_underline = "single"
    else:
        s_underline = "none"
    cell.font = Font(name=s_font, size=str(i_font_size), bold=b_bold, italic=b_italic, underline=s_underline)
    if fill_color is not None:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if align is not None:
        cell.alignment = Alignment(horizontal=align)
    if s_format is not None:
        cell.number_format = s_format
    if s_border is not None:
        cell.border = s_border


def create_package_summary_xls(l_package_info: List):
    """!
    @brief Update xls fiel with package summary
    @param l_package_info : list with package infos
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PackageInfo"
    for i, title in enumerate(L_PACKAGE_INFO_TITLE):
        set_cell(worksheet, 1, i + 1, title, b_bold=True)
    l_added_package = []
    i_package = 0
    for package in l_package_info:
        if package not in l_added_package:
            l_added_package.append(package)
            for i, value in enumerate(package):
                if isinstance(value, list):
                    if not value:
                        value = ""
                    else:
                        value = "\n".join(value)
                elif isinstance(value, str):
                    pass
                elif not value:
                    value = ""
                set_cell(worksheet, i_package + 2, i + 1, str(value))
            i_package += 1
    table_style = table.TableStyleInfo(name="TableStyleLight15",
                                       showRowStripes=True)
    s_end_cell = get_column_letter(len(l_added_package[0])) + str(len(l_added_package) + 1)
    report_table = table.Table(ref=f"A1:{s_end_cell}",
                               displayName=worksheet.title,
                               tableStyleInfo=table_style)
    worksheet.add_table(report_table)
    set_column_autowidth(worksheet)
    worksheet.freeze_panes = "C2"
    workbook.save(filename="PackageInfos.xlsx")
    log.info("Write %s package infos to file", len(l_added_package))


def get_package_info(package: str) -> PackageInfo:
    """!
    @brief Upgrade package from text file to latest version
    @param  package : check latest version of this package
    @return package info
    """
    try:
        url = f"https://pypi.org/pypi/{package}/json"
        response = requests.get(url, timeout=I_TIMEOUT)
        response.raise_for_status()
        package_info = response.json()
    except requests.exceptions.RequestException as e:
        log.error("Error occurred: %s", e)
        package_info = None
    else:
        l_requires_dist = []
        requires_dist = package_info["info"]["requires_dist"]
        if requires_dist:
            for req in requires_dist:
                l_requires_dist.append(req)
        package_info = PackageInfo(package,
                                   package_info["info"]["version"],
                                   package_info["info"]["author"],
                                   package_info["info"]["author_email"],
                                   package_info["info"]["license"],
                                   package_info["info"]["home_page"],
                                   package_info["info"]["package_url"],
                                   l_requires_dist,
                                   package_info["info"]["requires_python"])
    return package_info


def update_packages(filename: str) -> List:
    """!
    @brief  Update package from text file to latest version
    @param  filename : file name
    @return  list with updated packages
    """
    with open(filename, mode="r", encoding="utf-8") as file:
        lines = file.readlines()

    i_update_cnt = 0
    l_package_info = []
    updated_lines = []
    for line in lines:
        if re.match(r"^\s*#", line):  # comment line
            updated_lines.append(line)
        elif re.match(r"^\s*([\w\-]+)==([\w\.\-]+)\s*(#.*)?$", line):  # check for fix version 'package==version'
            package, version, comment = re.findall(r"^\s*([\w\-]+)==([\w\.\-]+)\s*(#.*)?$", line)[0]
            if comment != "":
                comment = f" {comment}"
            package_info = get_package_info(package)
            if (package_info is not None) and (package not in L_IGNORE_PACKAGES):
                if package_info.version and (package_info.version != version):
                    updated_lines.append(f"{package}=={package_info.version}{comment}\n")
                    i_update_cnt += 1
                    log.info("Updated: %s %s", package, package_info.version)
                else:
                    updated_lines.append(line)
                    log.debug(line)
                l_package_info.append(package_info)
            else:
                updated_lines.append(line)
                log.debug(line)
        else:  # package without version
            updated_lines.append(line)
            if line.strip():
                package_info = get_package_info(line)
                if package_info is not None:
                    l_package_info.append(package_info)
                    log.debug(line)

    with open(filename, mode="w", encoding="utf-8") as file:
        file.writelines(updated_lines)
        log.info("%s packages updates in %s\n", i_update_cnt, filename)
    return l_package_info


if __name__ == "__main__":
    l_all_package_info = []
    for req_file in L_FILES:
        l_file_package = update_packages(req_file)
        l_all_package_info += l_file_package
    if l_all_package_info:
        create_package_summary_xls(l_all_package_info)
    log.info("All packages checked and updated!")
