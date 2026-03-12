"""!
********************************************************************************
@file   openpyxl_util.py
@brief  Utility classes and constants for creating styled XLSX files with openpyxl.
********************************************************************************
"""

from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import table
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

FONT_SIZE = 11
FONT_NAME = "Calibri"
MAX_COLUMN_WIDTH = 30
MAX_ROW_HEIGHT = 1.3

# https://www.farb-tabelle.de/de/farbtabelle.htm
COLOR_BLACK = "000000"
COLOR_BLUE = "0000FF"
COLOR_BROWN = "A52A2A"
COLOR_GREY = "BEBEBE"
COLOR_GREEN = "00FF00"
COLOR_ORANGE = "FFA500"
COLOR_RED = "FF0000"
COLOR_VIOLET = "EE82EE"
COLOR_WHITE = "FFFFFF"
COLOR_YELLOW = "FFFF00"

THIN_BORDER = Border(left=Side(style="thin"),
                     right=Side(style="thin"),
                     top=Side(style="thin"),
                     bottom=Side(style="thin"))

NUMBER_FORMAT_CURRENCY = "#,##0.00"
NUMBER_FORMAT_EUR = "#,##0.00 €"
NUMBER_FORMAT_PERCENT = "0%"
NUMBER_FORMAT_DATETIME = "YYYY-MM-DD HH:MM:SS"
NUMBER_FORMAT_TIME = "hh:mm"

PAGE_MARGIN_FACTOR = 1 / 2.54  # convert cm to inch


class XLSCreator:
    """!
    @brief XLSX file creator with cell styling, auto-sizing and table formatting support.
    @param font_name : default font name
    @param font_size : default font size
    """

    def __init__(self, font_name: str | None = None, font_size: int | None = None) -> None:
        self.font_name = FONT_NAME if (font_name is None) else font_name
        self.font_size = FONT_SIZE if (font_size is None) else font_size
        self.workbook = Workbook()

    def save(self, filename: str) -> None:
        """!
        @brief Save the workbook to an XLSX file.
        @param filename : output xlsx filename
        """
        self.workbook.save(filename=filename)

    def set_table(self, worksheet: Worksheet, max_col: int, max_row: int, min_col: int = 1, min_row: int = 1) -> None:
        """!
        @brief Add a styled table to the worksheet range.
        @param worksheet : target worksheet to modify
        @param max_col : maximum column count for the table range
        @param max_row : maximum row count for the table range
        @param min_col : starting column for the table range, 1-based (default is first column)
        @param min_row : starting row for the table range, 1-based (default is first row)
        """
        table_style = table.TableStyleInfo(name="TableStyleLight15",
                                           showRowStripes=True)
        start_cell = get_column_letter(min_col) + str(min_row)
        end_cell = get_column_letter(max_col) + str(max_row)
        new_table = table.Table(ref=f"{start_cell}:{end_cell}",
                                displayName=worksheet.title,
                                tableStyleInfo=table_style)
        worksheet.add_table(new_table)

    def set_page_margins(self, worksheet: Worksheet, left: float | None = None, right: float | None = None,
                         top: float | None = None, bottom: float | None = None) -> None:
        """!
        @brief Set page margins in cm.
        @param worksheet : target worksheet to modify
        @param left : left margin value in cm
        @param right : right margin value in cm
        @param top : top margin value in cm
        @param bottom : bottom margin value in cm
        """
        if left is not None:
            worksheet.page_margins.left = left * PAGE_MARGIN_FACTOR
        if right is not None:
            worksheet.page_margins.right = right * PAGE_MARGIN_FACTOR
        if top is not None:
            worksheet.page_margins.top = top * PAGE_MARGIN_FACTOR
        if bottom is not None:
            worksheet.page_margins.bottom = bottom * PAGE_MARGIN_FACTOR

    def set_column_autowidth(self, worksheet: Worksheet, limited: bool = True) -> None:
        """!
        @brief Set automatic column width of worksheet.
        @param worksheet : target worksheet to modify
        @param limited : True: limit to MAX_COLUMN_WIDTH, False: no limit
        """
        for i, col_cells in enumerate(worksheet.columns, start=1):
            max_col_len = 0
            for j, cell in enumerate(col_cells):
                if limited:
                    if j != 0 and cell.value is not None:  # do not use first line description
                        max_col_len = max(max_col_len, len(str(cell.value).split("\n", maxsplit=1)[0]))
                else:
                    for line in str(cell.value).split("\n") if cell.value is not None else []:
                        max_col_len = max(max_col_len, len(line))
            if limited:
                max_col_len = min(max_col_len, MAX_COLUMN_WIDTH)
            if max_col_len == 0:
                worksheet.column_dimensions[get_column_letter(i)].hidden = True  # hide empty columns
            else:
                worksheet.column_dimensions[get_column_letter(i)].width = (max_col_len + 1) * 0.10 * self.font_size

    def set_row_autoheight(self, worksheet: Worksheet, limited: bool = True) -> None:
        """!
        @brief Set automatic row height of worksheet.
        @param worksheet : target worksheet to modify
        @param limited : True: use default row height, False: auto-height based on content
        """
        for i, row_cells in enumerate(worksheet.rows, start=1):
            row_height = MAX_ROW_HEIGHT
            if not limited:
                for cell in row_cells:
                    if cell.value is not None:
                        row_height = max(row_height, len(str(cell.value).split("\n")) * MAX_ROW_HEIGHT)
            worksheet.row_dimensions[i].height = row_height * self.font_size

    def set_cell(self, ws: Worksheet, row: int, column: int, value: str | int | float | datetime | time | None = None,
                 font_name: str | None = None, color: str | None = None, font_size: int | None = None,
                 bold: bool = False, italic: bool = False, strike: bool = False, underline: str | None = None,
                 vert_align: str | None = None, fill_color: str | None = None,
                 align: str | None = None, align_vert: str | None = "center", wrap_text: bool = False,
                 number_format: str | None = None, border: Border | None = None) -> None:
        """!
        @brief Set cell value with font, alignment, fill and border styling
               Font: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.fonts.html
               PatternFill: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.fills.html#openpyxl.styles.fills.PatternFill
               Alignment: https://openpyxl.readthedocs.io/en/latest/api/openpyxl.styles.alignment.html
        @param ws : target worksheet
        @param row : row index (1-based)
        @param column : column index (1-based)
        @param value : cell value [numeric, time, string, bool, None]
        @param font_name : font name (default is Calibri)
        @param color : font color e.g. "FF0000" for red
        @param font_size : font size
        @param bold : status if cell content should be bold
        @param italic : status if cell content should be italic
        @param strike : True to apply strikethrough formatting
        @param underline : underline options ['double', 'doubleAccounting', 'single', 'singleAccounting']
        @param vert_align : vertical align options ['superscript', 'subscript', 'baseline']
        @param fill_color : background fill color of cell
        @param align : horizontal text align option ["general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"]
        @param align_vert : vertical text align of cell ["top", "center", "bottom", "justify", "distributed"]
        @param wrap_text : wrap text
        @param number_format : number format of cell
        @param border : border of cell
        """
        if font_name is None:
            font_name = self.font_name
        if font_size is None:
            font_size = self.font_size
        cell = ws.cell(row=row, column=column, value=value)
        cell.font = Font(name=font_name, color=color, size=font_size, bold=bold, italic=italic,
                         strikethrough=strike, underline=underline, vertAlign=vert_align)
        if fill_color is not None:
            cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
        cell.alignment = Alignment(horizontal=align, vertical=align_vert, wrap_text=wrap_text)
        if number_format is not None:
            cell.number_format = number_format
        if border is not None:
            cell.border = border
